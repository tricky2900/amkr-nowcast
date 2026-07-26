"""Build output/amkr_indicator_data.xlsx for the Excel model to link against.

CONTRACT WITH YOUR MODEL: sheet names and column headers below are stable and
will not change between refreshes. Only rows get appended. Point Power Query at
these sheets (or at the CSVs in data/) and the connection never needs rebuilding.
Do not edit this workbook by hand - it is overwritten on every run. Put your own
work in a separate workbook that references it.
"""
from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from common import DATA, OUTPUT, log, safe_read

LOG = log("make_excel")

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
BODY = Font(name=ARIAL, size=10)
BLUE = Font(name=ARIAL, size=10, color="0000FF")     # hardcoded input
NOTE = Font(name=ARIAL, size=9, italic=True, color="595959")

MONEY = '$#,##0;($#,##0);-'
PCT = '0.0%;(0.0%);-'
NUM = '#,##0.0;(#,##0.0);-'


def _write(ws, df: pd.DataFrame, pct_cols=(), money_cols=(), num_cols=()):
    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for rec in df.itertuples(index=False):
        ws.append(list(rec))
    for j, col in enumerate(df.columns, start=1):
        letter = get_column_letter(j)
        fmt = PCT if col in pct_cols else MONEY if col in money_cols else NUM if col in num_cols else None
        width = max(11, min(30, int(df[col].astype(str).str.len().max() if len(df) else 10) + 3,
                            len(str(col)) + 4) if len(df) else 14)
        ws.column_dimensions[letter].width = max(width, len(str(col)) + 3)
        for i in range(2, len(df) + 2):
            cell = ws.cell(row=i, column=j)
            cell.font = BODY
            if fmt:
                cell.number_format = fmt
    ws.freeze_panes = "A2"


def run() -> str:
    OUTPUT.mkdir(exist_ok=True)
    wb = Workbook()

    # ---- legend ---------------------------------------------------------
    ws = wb.active
    ws.title = "README"
    lines = [
        ("AMKR Nowcast - Indicator Database", True),
        ("", False),
        ("This workbook is REGENERATED on every refresh. Do not edit it by hand;", False),
        ("link to it from your own model instead. Sheet names and column headers", False),
        ("are stable, so a Power Query connection built once keeps working.", False),
        ("", False),
        ("Sheets", True),
        ("Data_Monthly    Tidy long format: one row per period + series. Best for Power Query.", False),
        ("Data_Wide       Monthly levels pivoted, one column per series. Best for eyeballing.", False),
        ("Data_Quarterly  Calendar-quarter aggregates with YoY and QoQ. Complete quarters only.", False),
        ("QTD             The quarter in progress, months reported so far. Nowcast input.", False),
        ("Lead_Lag        Best correlation and implied lead for each series vs AMKR YoY.", False),
        ("Scenarios       Guidance-anchored bull / base / bear for the current quarter.", False),
        ("", False),
        ("Units", True),
        ("Taiwan company revenue   TWD thousands (as filed)", False),
        ("SIA billings             USD millions - NOTE: 3-month moving average, not discrete months", False),
        ("Taiwan exports           USD millions", False),
        ("AMKR revenue             USD millions", False),
        ("OSAT_COMPOSITE           Index, first available month = 100, revenue-weighted", False),
        ("", False),
        ("Health warnings", True),
        ("- All correlation work is in YoY space. Level correlations here are ~0.95+ and meaningless.", False),
        ("- 5 years of history is 20 quarters. Small sample; treat coefficients as indicative.", False),
        ("- SIA data is smoothed and lags turning points. Confirmation, not a leading signal.", False),
        ("- ASE monthly revenue includes USI's EMS business, which dilutes the pure OSAT read.", False),
        ("- ChipMOS and PTI skew memory/display; AMKR skews advanced SiP, comms and auto.", False),
        ("  Divergence between them is informative, not necessarily an error.", False),
    ]
    for i, (text, is_hdr) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(name=ARIAL, size=11, bold=True) if is_hdr else Font(name=ARIAL, size=10)
    ws.column_dimensions["A"].width = 100

    # ---- data sheets ----------------------------------------------------
    monthly = safe_read(DATA / "master_monthly.csv", required=True)
    _write(wb.create_sheet("Data_Monthly"), monthly, num_cols=("value",))

    wide = safe_read(DATA / "master_wide.csv")
    if not wide.empty:
        _write(wb.create_sheet("Data_Wide"), wide,
               num_cols=[c for c in wide.columns if c != "period"])

    quarterly = safe_read(DATA / "master_quarterly.csv")
    if not quarterly.empty:
        _write(wb.create_sheet("Data_Quarterly"), quarterly,
               pct_cols=("yoy", "qoq"), num_cols=("value",))

    qtd = safe_read(DATA / "qtd.csv")
    if not qtd.empty:
        _write(wb.create_sheet("QTD"), qtd, pct_cols=("qtd_yoy",), num_cols=("qtd_value",))

    ll = safe_read(OUTPUT / "lead_lag.csv")
    if not ll.empty:
        _write(wb.create_sheet("Lead_Lag"), ll, num_cols=("corr", "r_squared"))

    # ---- scenarios, with live formulas ----------------------------------
    sc = safe_read(OUTPUT / "scenarios.csv")
    if not sc.empty:
        ws = wb.create_sheet("Scenarios")
        _write(ws, sc,
               pct_cols=("predictor_qtd_yoy", "implied_amkr_yoy", "delta_vs_guide_mid_pct",
                         "model_r2", "resid_sd_pp"),
               money_cols=("prior_year_revenue_usdm", "bear_usdm", "base_usdm", "bull_usdm",
                           "guide_low_usdm", "guide_mid_usdm", "guide_high_usdm",
                           "delta_vs_guide_mid_usdm"))
        cols = {c: i + 1 for i, c in enumerate(sc.columns)}
        # Recompute the guidance deltas as formulas so overriding a guidance
        # cell updates the comparison instead of leaving a stale number.
        for r in range(2, len(sc) + 2):
            base = f"{get_column_letter(cols['base_usdm'])}{r}"
            mid = f"{get_column_letter(cols['guide_mid_usdm'])}{r}"
            ws.cell(row=r, column=cols["delta_vs_guide_mid_usdm"],
                    value=f"=IFERROR({base}-{mid},\"\")").number_format = MONEY
            ws.cell(row=r, column=cols["delta_vs_guide_mid_pct"],
                    value=f"=IFERROR({base}/{mid}-1,\"\")").number_format = PCT
            # Guidance is a hand-entered input: mark it blue per model convention.
            for k in ("guide_low_usdm", "guide_mid_usdm", "guide_high_usdm"):
                ws.cell(row=r, column=cols[k]).font = BLUE
        n = len(sc) + 3
        ws.cell(row=n, column=1, value=(
            "Blue cells are hand-entered from Amkor's 8-K guidance "
            "(data/manual/amkr_guidance.csv). Bear/bull are the wider of the "
            "guidance range and the model band, so the range never implies more "
            "precision than the company's own outlook."
        )).font = NOTE

    path = OUTPUT / "amkr_indicator_data.xlsx"
    wb.save(path)
    LOG.info("wrote %s (%d sheets)", path, len(wb.sheetnames))
    return str(path)


if __name__ == "__main__":
    run()
